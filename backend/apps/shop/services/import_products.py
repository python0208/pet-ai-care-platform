import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image as PILImage

from apps.shop.models import (
    Product,
    ProductCategory,
    ProductImportBatch,
    ProductImportRow,
    ProductInventory,
)


HEADER_ALIASES = {
    "图片": "image",
    "名称": "name",
    "单位": "unit",
    "进货价": "purchase_price",
    "规格": "spec",
    "零售价": "retail_price",
    "条码": "barcode",
    "重量": "weight",
    "直营店序号": "store_code",
    "分类": "category",
    "保质期（月）": "shelf_life_months",
    "保质期(月)": "shelf_life_months",
    "当前库存": "stock_quantity",
}

REQUIRED_FIELDS = {
    "name": "名称",
    "barcode": "条码",
    "retail_price": "零售价",
    "category": "分类",
    "stock_quantity": "当前库存",
}


@dataclass
class RowImportResult:
    status: str
    product: Product | None = None
    error_message: str = ""
    image_saved: bool = False
    image_failed: bool = False


def normalize_header(value):
    return str(value or "").strip().replace("（", "(").replace("）", ")")


def normalize_barcode(value):
    if value is None:
        return ""
    if isinstance(value, str):
        cleaned = value.strip()
    else:
        cleaned = str(value).strip()
    if not cleaned:
        return ""
    try:
        decimal_value = Decimal(cleaned)
    except InvalidOperation:
        return cleaned
    if decimal_value == decimal_value.to_integral_value():
        return str(decimal_value.quantize(Decimal("1")))
    normalized = format(decimal_value.normalize(), "f")
    return normalized.rstrip("0").rstrip(".")


def normalize_store_code(value):
    if value is None:
        return ProductInventory.DEFAULT_STORE_CODE
    cleaned = str(value).strip()
    return cleaned or ProductInventory.DEFAULT_STORE_CODE


def parse_decimal(value, field_name, required=False):
    if value is None or str(value).strip() == "":
        if required:
            raise ValueError(f"{field_name}不能为空")
        return None
    try:
        return Decimal(str(value).strip()).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field_name}格式不正确") from exc


def parse_int(value, field_name, required=False):
    if value is None or str(value).strip() == "":
        if required:
            raise ValueError(f"{field_name}不能为空")
        return None
    try:
        decimal_value = Decimal(str(value).strip())
    except InvalidOperation as exc:
        raise ValueError(f"{field_name}格式不正确") from exc
    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f"{field_name}必须为整数")
    if decimal_value < 0:
        raise ValueError(f"{field_name}不能为负数")
    return int(decimal_value)


def safe_filename_part(value):
    cleaned = re.sub(r"[^0-9A-Za-z_-]+", "_", str(value or "").strip())
    return cleaned.strip("_") or "product"


def extract_images_by_row(worksheet):
    images_by_row = {}
    for image in getattr(worksheet, "_images", []):
        try:
            row_number = image.anchor._from.row + 1
            images_by_row[row_number] = image
        except Exception:
            continue
    return images_by_row


def get_image_bytes(image):
    data = image._data()
    return data if isinstance(data, bytes) else bytes(data)


def save_product_image(image_or_bytes, barcode="", row_number=None):
    if isinstance(image_or_bytes, bytes):
        image_bytes = image_or_bytes
    else:
        image_bytes = get_image_bytes(image_or_bytes)

    image = PILImage.open(BytesIO(image_bytes))
    image_format = (image.format or "PNG").lower()
    if image_format == "jpeg":
        extension = "jpg"
    elif image_format in {"png", "webp", "jpg"}:
        extension = image_format
    else:
        extension = "png"

    filename_base = safe_filename_part(barcode) if barcode else f"row_{row_number or 'unknown'}"
    relative_path = f"products/{filename_base}.{extension}"
    if default_storage.exists(relative_path):
        default_storage.delete(relative_path)
    default_storage.save(relative_path, ContentFile(image_bytes))
    return f"{settings.MEDIA_URL.rstrip('/')}/{relative_path}"


class ProductExcelImporter:
    def __init__(self, batch: ProductImportBatch):
        self.batch = batch
        self.images_by_row = {}
        self.header_map = {}

    def import_file(self):
        try:
            workbook = load_workbook(self.batch.file.path)
            worksheet = workbook.active
            self.images_by_row = extract_images_by_row(worksheet)
            self.header_map = self.build_header_map(worksheet)
            self.import_worksheet(worksheet)
        except Exception as exc:
            self.batch.status = ProductImportBatch.Status.FAILED
            self.batch.error_message = str(exc)
            self.batch.completed_at = timezone.now()
            self.batch.save(
                update_fields=[
                    "status",
                    "error_message",
                    "completed_at",
                ]
            )
            return self.batch
        return self.batch

    def build_header_map(self, worksheet):
        header_map = {}
        for column_index, cell in enumerate(worksheet[1], start=1):
            header = normalize_header(cell.value)
            field_name = HEADER_ALIASES.get(header)
            if field_name:
                header_map[field_name] = column_index
        missing = [
            label
            for field_name, label in REQUIRED_FIELDS.items()
            if field_name not in header_map
        ]
        if missing:
            raise ValueError(f"缺少必要表头：{', '.join(missing)}")
        return header_map

    def import_worksheet(self, worksheet):
        total_rows = 0
        success_count = 0
        updated_count = 0
        failed_count = 0
        image_success_count = 0
        image_failed_count = 0

        for row_number in range(2, worksheet.max_row + 1):
            row_values = self.get_row_values(worksheet, row_number)
            if self.is_empty_row(row_values):
                continue
            total_rows += 1
            result = self.import_row(row_number, row_values)
            if result.status == ProductImportRow.Status.SUCCESS:
                success_count += 1
            elif result.status == ProductImportRow.Status.UPDATED:
                updated_count += 1
            else:
                failed_count += 1
            if result.image_saved:
                image_success_count += 1
            if result.image_failed:
                image_failed_count += 1

        self.batch.total_rows = total_rows
        self.batch.success_count = success_count
        self.batch.updated_count = updated_count
        self.batch.failed_count = failed_count
        self.batch.image_success_count = image_success_count
        self.batch.image_failed_count = image_failed_count
        if failed_count and (success_count or updated_count):
            self.batch.status = ProductImportBatch.Status.PARTIAL_SUCCESS
        elif failed_count:
            self.batch.status = ProductImportBatch.Status.FAILED
        else:
            self.batch.status = ProductImportBatch.Status.SUCCESS
        self.batch.completed_at = timezone.now()
        self.batch.save()

    def get_row_values(self, worksheet, row_number):
        values = {}
        for field_name, column_index in self.header_map.items():
            values[field_name] = worksheet.cell(row=row_number, column=column_index).value
        return values

    def is_empty_row(self, row_values):
        return all(value is None or str(value).strip() == "" for value in row_values.values())

    def import_row(self, row_number, row_values):
        product_name = str(row_values.get("name") or "").strip()
        barcode = normalize_barcode(row_values.get("barcode"))
        try:
            with transaction.atomic():
                result = self.save_row(row_number, row_values, product_name, barcode)
                ProductImportRow.objects.create(
                    batch=self.batch,
                    row_number=row_number,
                    product_name=product_name,
                    barcode=barcode,
                    status=result.status,
                    error_message=result.error_message,
                    product=result.product,
                )
                return result
        except Exception as exc:
            ProductImportRow.objects.create(
                batch=self.batch,
                row_number=row_number,
                product_name=product_name,
                barcode=barcode,
                status=ProductImportRow.Status.FAILED,
                error_message=str(exc),
            )
            return RowImportResult(status=ProductImportRow.Status.FAILED, error_message=str(exc))

    def save_row(self, row_number, row_values, product_name, barcode):
        if not product_name:
            raise ValueError("名称不能为空")
        if not barcode:
            raise ValueError("条码不能为空")
        category_name = str(row_values.get("category") or "").strip()
        if not category_name:
            raise ValueError("分类不能为空")

        retail_price = parse_decimal(row_values.get("retail_price"), "零售价", required=True)
        purchase_price = parse_decimal(row_values.get("purchase_price"), "进货价")
        weight = parse_decimal(row_values.get("weight"), "重量")
        shelf_life_months = parse_int(row_values.get("shelf_life_months"), "保质期（月）")
        stock_quantity = parse_int(row_values.get("stock_quantity"), "当前库存", required=True)
        store_code = normalize_store_code(row_values.get("store_code"))

        category, _ = ProductCategory.objects.get_or_create(name=category_name)
        product, created = Product.objects.get_or_create(
            barcode=barcode,
            defaults={
                "name": product_name,
                "unit": str(row_values.get("unit") or "").strip(),
                "spec": str(row_values.get("spec") or "").strip(),
                "category": category,
                "purchase_price": purchase_price,
                "retail_price": retail_price,
                "weight": weight,
                "shelf_life_months": shelf_life_months,
                "status": Product.Status.ACTIVE,
            },
        )
        if not created:
            product.name = product_name
            product.unit = str(row_values.get("unit") or "").strip()
            product.spec = str(row_values.get("spec") or "").strip()
            product.category = category
            product.purchase_price = purchase_price
            product.retail_price = retail_price
            product.weight = weight
            product.shelf_life_months = shelf_life_months
            product.status = Product.Status.ACTIVE

        image_saved = False
        image_failed = False
        warnings = []
        image = self.images_by_row.get(row_number)
        if image:
            try:
                product.cover_image = save_product_image(image, barcode=barcode, row_number=row_number)
                image_saved = True
            except Exception as exc:
                image_failed = True
                warnings.append(f"图片提取失败：{exc}")

        product.save()

        ProductInventory.objects.update_or_create(
            product=product,
            store_code=store_code,
            defaults={
                "stock_quantity": stock_quantity,
                "last_imported_at": timezone.now(),
            },
        )
        return RowImportResult(
            status=ProductImportRow.Status.SUCCESS if created else ProductImportRow.Status.UPDATED,
            product=product,
            error_message="；".join(warnings),
            image_saved=image_saved,
            image_failed=image_failed,
        )


def import_product_excel(batch):
    return ProductExcelImporter(batch).import_file()
